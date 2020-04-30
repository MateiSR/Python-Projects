from openpyxl import workbook, load_workbook

workbook = load_workbook(filename = "workbook.xlsx")
print(workbook.sheetnames)
sheet = workbook.active
print(sheet["A1"].value)