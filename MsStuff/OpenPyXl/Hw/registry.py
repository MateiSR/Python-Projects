import json
from openpyxl import workbook, load_workbook

# Load workbook 
workbook = load_workbook(filename="registry.xlsx")
sheet = workbook.active

registry = {}

# Cycle through all the rows
for row in sheet.iter_rows(min_row = 2, min_col = 0, values_only=True):
    current_id = row[0]
    current = {
        "doctor": row[1],
        "patient": row[2],
        "nurses": row[3],
        "health": row[4],
        "days_in": row[5]
    }
    registry[current_id] = current

# Dump and format to json    
    with open('output.json', 'w') as fp:
        json.dump(registry, fp, sort_keys=True, indent=4)